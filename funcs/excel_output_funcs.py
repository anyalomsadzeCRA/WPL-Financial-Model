import pandas as pd
import numpy as np
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from openpyxl.styles import numbers

### Formatting Functions

# Function to format a DataFrame with currency values
def style_dataframe_with_currency(df):
    """
    Styles DataFrame with currency format.

    Parameters:
        df (DataFrame): Input DataFrame.

    Returns:
        DataFrame: Styled DataFrame with currency format.
    """
    styled_df = df.copy()
    styled_df = styled_df.applymap(lambda x: '${:,.0f}'.format(x) if pd.notna(x) and np.issubdtype(type(x), np.number) else x)
    return styled_df

# Function to apply color fill to a row in a worksheet
def apply_color_fill_to_row(ws, row_number, start_column, end_column, color="ABAAAA"):
    """
    Applies color fill to a row in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.
        start_column (int): Starting column number.
        end_column (int): Ending column number.
        color (str): Color code for fill.

    Returns:
        None
    """
    for col_num in range(start_column, end_column + 1):
        cell = ws.cell(row=row_number, column=col_num)
        cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        
# Function to apply text color to a row in a worksheet
def apply_text_color_to_row(ws, row_number, start_column, end_column, color="000000"):
    """
    Applies text color to a row in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.
        start_column (int): Starting column number.
        end_column (int): Ending column number.
        color (str): Color code for text.

    Returns:
        None
    """
    for col_num in range(start_column, end_column + 1):
        cell = ws.cell(row=row_number, column=col_num)
        cell.font = Font(color=color)
        
# Function to apply bold font to a row in a worksheet
def apply_bold_to_row(ws, row_number, start_column, end_column):
    """
    Applies bold font to a row in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.
        start_column (int): Starting column number.
        end_column (int): Ending column number.

    Returns:
        None
    """
    for col_num in range(start_column, end_column + 1):
        cell = ws.cell(row=row_number, column=col_num)
        cell.font = Font(bold=True)

# Function to apply alignment to a row in a worksheet
def apply_alignment_to_row(ws, row_number, start_column, end_column):
    """
    Applies alignment to a row in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.
        start_column (int): Starting column number.
        end_column (int): Ending column number.

    Returns:
        None
    """
    for col_num in range(start_column, end_column + 1):
        cell = ws.cell(row=row_number, column=col_num)
        cell.alignment = Alignment(horizontal='center')
        
# Function to apply currency format to a cell in a worksheet
def apply_currency_format(ws, row_number, column_number):
    """
    Applies currency format to a cell in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.
        column_number (int): Column number.

    Returns:
        None
    """
    cell = ws.cell(row=row_number, column=column_number)
    value = cell.value
    
    try:
        if value != '' and isinstance(value, str) and '$' in value:
            numeric_value = float(re.sub(r'[^\d.]', '', value))
            cell.value = numeric_value
            currency_format_code = numbers.BUILTIN_FORMATS[44]  # Format code for currency
            cell.number_format = currency_format_code
            cell.alignment = Alignment(horizontal='center')
    except ValueError:
        cell.value = value
        
# Function to set column widths in a worksheet
def set_column_widths(ws):
    """
    Sets column widths in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.

    Returns:
        None
    """
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 25
    
    # Set the width for columns C through Z
    for col_letter in range(ord('C'), ord('Z')+1):
        col_letter = chr(col_letter)
        ws.column_dimensions[col_letter].width = 18

###### Text Formatting Functions

# Function to apply text formatting to rows in a worksheet
def apply_text_formatting_to_rows(ws, start_row, end_row):
    """
    Applies text formatting to rows in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        start_row (int): Starting row number.
        end_row (int): Ending row number.

    Returns:
        None
    """
    for row_number in range(start_row, end_row + 1):
        for cell in ws[row_number]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

# Function to apply formatting to a year row in a worksheet
def apply_formatting_to_year_row(ws, row_number):
    """
    Applies formatting to a year row in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        row_number (int): Row number.

    Returns:
        None
    """
    apply_bold_to_row(ws, row_number=row_number, start_column=1, end_column=100)
    apply_alignment_to_row(ws, row_number=row_number, start_column=1, end_column=100)
    apply_text_color_to_row(ws, row_number=row_number, start_column=1, end_column=100, color="0C0D0D")    
    
# Function to apply bold font and green color to monetary values in a worksheet
def apply_bold_text_green_money(ws, start_row, end_row):
    """
    Applies bold font and green color to monetary values in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        start_row (int): Starting row number.
        end_row (int): Ending row number.

    Returns:
        None
    """
    for row_number in range(start_row, end_row + 1):
        for col_num, cell in enumerate(ws[row_number], start=1):
            if cell.value is not None and cell.value != '':
                if col_num == 1:
                    # Format the first column
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                else:
                    # Format other columns
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                    cell.font = Font(bold=True, color="11AD11")  # Green text

# Function to apply blue color to monetary values in a worksheet
def apply_norm_text_blue_money(ws, start_row, end_row):
    """
    Applies blue color to monetary values in a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        start_row (int): Starting row number.
        end_row (int): Ending row number.

    Returns:
        None
    """
    for row_number in range(start_row, end_row + 1):
        for col_num, cell in enumerate(ws[row_number], start=1):
            if cell.value is not None and cell.value != '':
                if col_num != 1:
                    # Format other columns
                    cell.font = Font(color="0067A7")  # Blue text
                    cell.alignment = Alignment(horizontal='center')

###### Adding Data Functions    

# Function to add a header row to a worksheet
def add_header_row(ws, header_row, color="F2F2F2"):
    """
    Adds a header row to a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        header_row (list): List representing the header row.
        color (str): Color code for header row.

    Returns:
        None
    """
    # Find the first empty row in the worksheet
    next_empty_row = ws.max_row 
    
    # Add row
    ws.append(header_row)

    # Apply formatting
    if next_empty_row == 1:
        row_to_format = next_empty_row 
    else:
        row_to_format = next_empty_row + 1
    apply_color_fill_to_row(ws, row_number=row_to_format, start_column=1, end_column=100, color=color)
    apply_bold_to_row(ws, row_number=row_to_format, start_column=1, end_column=100)

# Function to add data to a worksheet
def add_data_to_worksheet(ws, df, 
                          formatting_type=None, 
                          use_cols_as_header=False, 
                          bold_last_row=False,
                          bold_first_row=False):
    """
    Adds data to a worksheet.

    Parameters:
        ws (Worksheet): Excel worksheet.
        df (DataFrame): DataFrame containing data to be added.
        formatting_type (str): Type of formatting to be applied.
        use_cols_as_header (bool): Whether to use DataFrame columns as header row.
        bold_last_row (bool): Whether to bold the last row.
        bold_first_row (bool): Whether to bold the first row.

    Returns:
        None
    """
    # Find the first empty row in the worksheet
    start_row_val = ws.max_row + 1
    curr_row_val = start_row_val

    # Add the column header row if needed
    if use_cols_as_header:
        header_row = list(df.columns)
        ws.append(header_row)
        curr_row_val += 1
    
    # Add the data rows
    for index, row in df.iterrows():
        ws.append(list(row))
        # Apply currency formatting only to cells with "$" sign in the row
        for col_num, value in enumerate(row, start=1):
            if isinstance(value, str) and '$' in value:
                apply_currency_format(ws, curr_row_val, col_num)
        curr_row_val += 1

    # Apply text formatting based on the specified type
    end_row_val = curr_row_val
    if formatting_type == "Bold Text":
        apply_text_formatting_to_rows(ws, start_row_val, end_row_val)
    elif formatting_type == "Bold Text and Green Money":
        apply_bold_text_green_money(ws, start_row_val, end_row_val)
    elif formatting_type == "Normal Text and Blue Money":
        apply_norm_text_blue_money(ws, start_row_val, end_row_val)
        
    # Bold the last row if specified
    if bold_last_row:
        apply_bold_to_row(ws, row_number=end_row_val-1, start_column=1, end_column=len(df.columns))
    if bold_first_row:
        if use_cols_as_header:
            apply_bold_to_row(ws, row_number=start_row_val+1, start_column=1, end_column=len(df.columns))
        else:
            apply_bold_to_row(ws, row_number=start_row_val, start_column=1, end_column=len(df.columns))
        
     # Apply formatting to column header
    if use_cols_as_header:
        apply_bold_to_row(ws, row_number=start_row_val, start_column=1, end_column=100)

# Function to display alternating dictionary dataframes with headers in a worksheet
def display_alternating_dict_dataframes_with_headers(worksheet, book_dict, tax_dict, header_text, subhead_color):
    """
    Displays alternating dictionary dataframes with headers in a worksheet.

    Parameters:
        worksheet (Worksheet): Excel worksheet.
        book_dict (dict): Dictionary containing book values.
        tax_dict (dict): Dictionary containing tax values.
        header_text (str): Header text.
        subhead_color (str): Color code for subheader.

    Returns:
        None
    """
    book_keys = list(book_dict.keys())
    tax_keys = list(tax_dict.keys())

    for book_key, tax_key in zip(book_keys, tax_keys):
        
        # Display subheader with the book depreciation value name
        subheader_row_book = [book_key]
        add_header_row(worksheet, subheader_row_book, color=subhead_color)

        book_value = book_dict[book_key]
        if isinstance(book_value, pd.DataFrame):
            book_value_styled = style_dataframe_with_currency(book_value).reset_index().rename(columns={'index': 'Book Depreciation'})
            add_data_to_worksheet(worksheet,
                                  book_value_styled,
                                  use_cols_as_header=True,
                                  formatting_type="Normal Text and Blue Money",
                                  bold_last_row=True)
        else:
            # If not a DataFrame, create a single-row DataFrame with the string value
            single_row_df = pd.DataFrame([book_value], columns=['Component'])
            add_data_to_worksheet(worksheet,
                                  single_row_df,
                                  use_cols_as_header=True,
                                  formatting_type="Normal Text and Blue Money",
                                  bold_last_row=True)

        tax_value = tax_dict[tax_key]
        if isinstance(tax_value, pd.DataFrame):
            tax_value_styled = style_dataframe_with_currency(tax_value).reset_index().rename(columns={'index': 'Tax Depreciation'})
            add_data_to_worksheet(worksheet,
                                  tax_value_styled,
                                  use_cols_as_header=True,
                                  formatting_type="Normal Text and Blue Money",
                                  bold_last_row=True)
        else:
            # If not a DataFrame, create a single-row DataFrame with the string value
            single_row_df = pd.DataFrame([tax_value], columns=['Component'])
            add_data_to_worksheet(worksheet,
                                  single_row_df,
                                  use_cols_as_header=True,
                                  formatting_type="Normal Text and Blue Money",
                                  bold_last_row=True)
